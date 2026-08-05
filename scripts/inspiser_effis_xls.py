"""Undersøker EFFIS' nedlastbare landtotaler før K3s parser skrives.

Filen ligger bak en 301 til ``data.effis.emergency.copernicus.eu``, som
utviklingsmiljøets nettverkspolicy avviser med 403 på CONNECT. Den kan derfor
ikke prøves derfra. Dette skriptet kjøres i GitHub Actions, som er samme sted
den månedlige ETL-kjøringen kjører.

Skriptet skal svare på to ting, og ikke mer:

1. Er filen hentbar programmatisk?
2. Hvor mange år dekker den?

Ingen parser skrives før svaret er rapportert.

Engangsverktøy. Slettes sammen med workflowen når K3 er koblet om.

Kjøres som modul fra repotoppen: ``python -m scripts.inspiser_effis_xls``
"""

import io
import urllib.error
import urllib.request

BRUKERAGENT = "skogbranner-etl/1.0 (+https://github.com/sheffielddivided/skogbranner)"

SIDE = "https://forest-fire.emergency.copernicus.eu/applications/data-and-services"

# Lenken slik den står på siden, merket «Country totals 2024».
XLS = (
    "https://forest-fire.emergency.copernicus.eu/effis/applications/"
    "data-and-services/report_2024.xlsx"
)

# Er filen knyttet til én årsrapport, finnes den kanskje per år. Det avgjør om
# K3 skal hente én fil eller flere.
ANDRE_AAR = [
    "https://forest-fire.emergency.copernicus.eu/effis/applications/"
    f"data-and-services/report_{aar}.xlsx"
    for aar in (2025, 2023, 2022, 2021, 2020)
]


def hent(url):
    forespoersel = urllib.request.Request(url, headers={"User-Agent": BRUKERAGENT})
    with urllib.request.urlopen(forespoersel, timeout=120) as svar:
        return svar.status, svar.geturl(), svar.headers.get("content-type", ""), svar.read()


def vis_arbeidsbok(data):
    """Skriver ut arkene, formen og årsspennet i filen."""
    import openpyxl

    bok = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    print(f"  ARK: {bok.sheetnames}")

    for navn in bok.sheetnames:
        ark = bok[navn]
        rader = [rad for rad in ark.iter_rows(values_only=True)]
        print(f"\n  --- {navn}: {len(rader)} rader x {ark.max_column} kolonner")
        for i, rad in enumerate(rader[:12]):
            celler = [str(c)[:24] if c is not None else None for c in rad[:12]]
            print(f"    {i + 1}: {celler}")
        if len(rader) > 12:
            print("    …")
            for i, rad in enumerate(rader[-4:], start=len(rader) - 3):
                celler = [str(c)[:24] if c is not None else None for c in rad[:12]]
                print(f"    {i}: {celler}")

        # Årstall hvor som helst i arket, så spennet kan leses av uten å vite
        # hvilken kolonne de står i.
        aar = set()
        for rad in rader:
            for celle in rad:
                try:
                    verdi = int(str(celle).strip())
                except (TypeError, ValueError):
                    continue
                if 1980 <= verdi <= 2030:
                    aar.add(verdi)
        if aar:
            print(f"    ÅRSTALL I ARKET: {min(aar)}–{max(aar)} ({len(aar)} ulike)")


def main():
    print("=" * 72)
    print("K3 EFFIS — er landtotalene hentbare, og hvor mange år dekker de?")
    print("=" * 72)

    print(f"\n### {XLS}")
    try:
        status, endelig, ctype, data = hent(XLS)
    except urllib.error.HTTPError as e:
        print(f"  IKKE HENTBAR: HTTP {e.code} {e.reason}")
        return
    except Exception as e:
        print(f"  IKKE HENTBAR: {type(e).__name__}: {e}")
        return

    print(f"  HTTP {status} | {ctype} | {len(data)} bytes")
    print(f"  endelig adresse: {endelig}")
    print(f"  magiske bytes:   {data[:8]!r}")

    if data[:2] == b"PK":
        print("  → ekte XLSX (zip-basert)")
        vis_arbeidsbok(data)
    elif data[:8] == b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1":
        print("  → gammel XLS (BIFF). Krever xlrd, ikke openpyxl.")
    else:
        print(f"  → ikke en arbeidsbok. Første 300 tegn: {data[:300]!r}")

    print("\n### Finnes filen for andre år?")
    for url in ANDRE_AAR:
        try:
            status, endelig, ctype, data = hent(url)
            print(f"  {url.rsplit('/', 1)[-1]}: HTTP {status} | {len(data)} bytes")
        except urllib.error.HTTPError as e:
            print(f"  {url.rsplit('/', 1)[-1]}: HTTP {e.code}")
        except Exception as e:
            print(f"  {url.rsplit('/', 1)[-1]}: {type(e).__name__}: {e}")


if __name__ == "__main__":
    main()
