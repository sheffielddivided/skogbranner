"""K3 — EFFIS: nasjonalt rapporterte landtotaler.

Tallene ligger som en nedlastbar XLS som EFFIS publiserer sammen med
årsrapporten «Forest fires in Europe, Middle East and North Africa». Filen har
to ark: brent areal i hektar, og antall skogbranner. Begge går tilbake til
1980.

Dette er **ikke** det samme som K4. K3 er tall landene selv har rapportert inn,
etter sine egne definisjoner. K4 er EFFIS' egen satellittkartlegging. Se
CLAUDE.md § 5.

Filnavnet bærer årsrapportens år, så katalogen prøves bakover og nyeste
tilgjengelige fil velges. Hver fil inneholder hele historikken fram til sitt
år, ikke bare det året.

Kilden oppgir hektar. Konverteringen til km² skjer i ``normalize.py``.

Kjøres som modul fra repotoppen: ``python -m etl.sources.k3_effis``
"""

import hashlib
import io
import json
import re
import urllib.error
import urllib.request
from datetime import date, datetime, timezone

import openpyxl

from etl.schema import RAW_DIR, SOURCES_JSON, STATUS_JSON

SOURCE_ID = "K3"
SERIES_BURNED_AREA = "effis_annual_country_totals"
SERIES_FIRE_COUNT = "effis_annual_country_fire_count"

# CLAUDE.md § 5: nasjonalt rapporterte tall, med hvert lands egne definisjoner.
KVALITET = "reported"

BASE_URL = (
    "https://forest-fire.emergency.copernicus.eu/effis/applications/"
    "data-and-services/report_{aar}.xlsx"
)
LANDING_URL = (
    "https://forest-fire.emergency.copernicus.eu/applications/data-and-services"
)
LICENSE_URL = "https://forest-fire.emergency.copernicus.eu/about-effis/data-license"

RAW_XLSX = RAW_DIR / "k3_effis_country_totals.xlsx"

# Ikke alle år har en fil — 2025 og 2022 svarer 404, 2024, 2023 og 2021 svarer
# 200. Nyeste fil finnes derfor ved å prøve bakover fra inneværende år framfor
# å låse et årstall.
MAKS_AAR_BAKOVER = 6

# Arknavnene er avkuttet av Excels grense på 31 tegn («Burnt area (ha) 1980 -
# 204»), så de matches på prefiks og ikke eksakt.
ARK_AREAL = "Burnt area"
ARK_ANTALL = "Nr. of forest fires"

# EFFIS bruker XKO for Kosovo. Vi bruker XKX — se CLAUDE.md § 6.
EFFIS_CODE_MAP = {
    "XKO": "XKX",
}

AARSTALL = re.compile(r"^\d{4}$")

BRUKERAGENT = "skogbranner-etl/1.0 (+https://github.com/sheffielddivided/skogbranner)"


def _hent(url):
    forespoersel = urllib.request.Request(url, headers={"User-Agent": BRUKERAGENT})
    with urllib.request.urlopen(forespoersel, timeout=180) as svar:
        return svar.read()


def _sha256(data):
    return hashlib.sha256(data).hexdigest()


def entity_kode(kode):
    """Oversetter EFFIS' landkode til vår entity-kode."""
    return EFFIS_CODE_MAP.get(kode, kode)


def finn_nyeste():
    """Prøver seg bakover fra inneværende år og henter den nyeste filen.

    Returnerer (årstall, innhold). Reiser feil hvis ingen av årene svarer — da
    har EFFIS lagt om adressen, og det skal stoppe kjøringen framfor å bli
    stående med et gammelt datasett uten at noen får vite det.
    """
    i_aar = date.today().year
    proevd = []
    for aar in range(i_aar, i_aar - MAKS_AAR_BAKOVER, -1):
        url = BASE_URL.format(aar=aar)
        try:
            data = _hent(url)
        except urllib.error.HTTPError as e:
            proevd.append(f"{aar}: HTTP {e.code}")
            continue
        # En 404-side kan komme som HTML med status 200. XLSX er zip-basert.
        if data[:2] != b"PK":
            proevd.append(f"{aar}: ikke en arbeidsbok")
            continue
        return aar, data

    raise ValueError(
        "fant ingen nedlastbar landtotalfil hos EFFIS. Prøvd: " + ", ".join(proevd)
    )


def _tall(celle):
    """Leser et tall slik regnearket skriver det.

    Formatet varierer mellom år: «44 251 » med hardt mellomrom som tusenskille
    og etterfølgende blank, «137651» uten, og «47711.13» med desimaler. Tomme
    celler er «ingen data», ikke null, og gir None.
    """
    if celle is None:
        return None
    if isinstance(celle, (int, float)):
        return float(celle)
    reint = str(celle).replace("\xa0", "").replace(" ", "").replace(",", "").strip()
    if not reint:
        return None
    try:
        return float(reint)
    except ValueError:
        return None


def _les_ark(bok, prefiks):
    """Leser ett ark til rader på formen (entity, år, verdi).

    Arket har årstall nedover og landkoder bortover. Tomme celler hoppes over —
    et land som ikke rapporterte et år, har ingen verdi, og skal ikke få 0.
    """
    navn = next((n for n in bok.sheetnames if n.startswith(prefiks)), None)
    if navn is None:
        raise ValueError(
            f"fant ingen ark som begynner med {prefiks!r}. Ark i filen: {bok.sheetnames}"
        )

    rader = [rad for rad in bok[navn].iter_rows(values_only=True)]
    if not rader:
        raise ValueError(f"arket {navn!r} er tomt")

    overskrifter = [str(c).strip() if c is not None else "" for c in rader[0]]

    ut = []
    for rad in rader[1:]:
        if not rad or rad[0] is None:
            continue
        aar_tekst = str(rad[0]).strip()
        if not AARSTALL.match(aar_tekst):
            continue
        aar = int(aar_tekst)

        for i, celle in enumerate(rad):
            if i == 0 or i >= len(overskrifter):
                continue
            kode = overskrifter[i]
            if not kode or kode.lower() == "year":
                continue
            verdi = _tall(celle)
            if verdi is None:
                continue
            ut.append({"code": kode, "year": aar, "value": verdi})

    if not ut:
        raise ValueError(f"arket {navn!r} ga ingen verdier")
    ut.sort(key=lambda r: (r["code"], r["year"]))
    return ut


def hent():
    """Laster ned nyeste landtotalfil, skriver den til data/raw/ og leser den.

    Returnerer (areal, antall, info). Arealet er hektar slik det står i kilden.
    """
    RAW_DIR.mkdir(parents=True, exist_ok=True)

    aar, data = finn_nyeste()
    RAW_XLSX.write_bytes(data)

    bok = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    areal = _les_ark(bok, ARK_AREAL)
    antall = _les_ark(bok, ARK_ANTALL)

    alle_aar = [r["year"] for r in areal] + [r["year"] for r in antall]
    info = {
        "source_id": SOURCE_ID,
        "series_id": [SERIES_BURNED_AREA, SERIES_FIRE_COUNT],
        "downloaded_at": date.today().isoformat(),
        "checksum": _sha256(data),
        "rows": len(areal) + len(antall),
        "report_year": aar,
        "countries": len({r["code"] for r in areal}),
        "coverage_start": min(alle_aar),
        "coverage_end": max(alle_aar),
    }
    return areal, antall, info


def skriv_metadata(info, processed_files):
    """Registrerer kilden i data/_sources.json.

    EFFIS har egen datalisens som må gjengis. Lenken til lisensteksten står i
    ``license_url`` og refereres i attribusjonsblokken — se CLAUDE.md § 5.
    """
    with open(SOURCES_JSON, encoding="utf-8") as f:
        sources = json.load(f)

    sources["sources"][SOURCE_ID] = {
        "source_id": SOURCE_ID,
        "name": "EFFIS — landtotaler, brent areal og antall skogbranner",
        "publisher": "Joint Research Centre, Europakommisjonen, under Copernicus",
        "url": LANDING_URL,
        "download_url": BASE_URL.format(aar=info["report_year"]),
        "license": "Creative Commons Attribution 4.0 International (CC BY 4.0). "
        "Copyright (C) European Union, 1995–2025. Kommisjonens gjenbrukspolitikk "
        "følger kommisjonsbeslutningen av 12. desember 2011 om gjenbruk av "
        "kommisjonsdokumenter.",
        "license_url": LICENSE_URL,
        "attribution": "European Forest Fire Information System (EFFIS), "
        "Copernicus Emergency Management Service, Joint Research Centre.",
        "requires_agreement": False,
        "geography": "Europa, Midtøsten og Nord-Afrika",
        "coverage_start": str(info["coverage_start"]),
        "coverage_end": str(info["coverage_end"]),
        "temporal_resolution": "annual",
        "quality": KVALITET,
        "unit_source": "hectares",
        "downloaded_at": info["downloaded_at"],
        "source_last_updated": str(info["report_year"]),
        "checksum": info["checksum"],
        "series": [SERIES_BURNED_AREA, SERIES_FIRE_COUNT],
        "processed_files": processed_files,
        "footnotes": ["f_reporting_basis", "f_coverage_change"],
        "notes": "Tallene er rapportert inn av det enkelte land og følger "
        "landets egne definisjoner av hva som telles som en skogbrann. De er "
        "hentet fra regnearket EFFIS publiserer sammen med årsrapporten. "
        "Serien begynner i 1980 med fem land, og flere kommer til utover i "
        "serien, så hvor langt tilbake tallene går, varierer mellom land. "
        "Dette er en annen kilde enn K4, som er EFFIS' egen satellittkartlegging.",
    }
    with open(SOURCES_JSON, "w", encoding="utf-8") as f:
        json.dump(sources, f, ensure_ascii=False, indent=2)
        f.write("\n")


def skriv_status(status, melding, info=None):
    """Skriver kjørestatus til data/_status.json."""
    with open(STATUS_JSON, encoding="utf-8") as f:
        data = json.load(f)

    naa = datetime.now(timezone.utc).isoformat(timespec="seconds")
    forrige = data["sources"].get(SOURCE_ID, {})
    data["last_run"] = naa
    data["sources"][SOURCE_ID] = {
        "status": status,
        "last_attempt": naa,
        "last_success": naa if status == "ok" else forrige.get("last_success"),
        "rows": info["rows"] if info else forrige.get("rows"),
        "checksum": info["checksum"] if info else forrige.get("checksum"),
        "message": melding,
    }
    with open(STATUS_JSON, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
        f.write("\n")


def main():
    areal, antall, info = hent()
    print(
        f"K3: rapport {info['report_year']}, {len(areal)} arealverdier og "
        f"{len(antall)} branntall for {info['countries']} land, "
        f"{info['coverage_start']}–{info['coverage_end']}"
    )
    return areal, antall, info


if __name__ == "__main__":
    main()
