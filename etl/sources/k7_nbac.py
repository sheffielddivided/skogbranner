"""K7 — CNFDB / NBAC: brent areal og antall branner i Canada.

To filer fra Canadian Wildland Fire Information System:

* **NBAC** gir brent areal per år som produsentens eget årsaggregat. Vi bruker
  aggregatet, ikke polygonene.
* **NFDB point stats** gir antall registrerte branner per år.

Filnavnet på NBAC-aggregatet bærer en utgivelsesdato, så katalogen leses og
den nyeste filen velges. Å låse et filnavn ville gjort at kilden stille sluttet
å oppdatere seg neste gang produsenten legger ut en ny utgave.

Kilden oppgir hektar. Konverteringen til km² skjer i ``normalize.py``.

Sluttbrukeravtalen er akseptert, og siteringen skal gjengis ordrett — se
CLAUDE.md § 5.

Kjøres som modul fra repotoppen: ``python -m etl.sources.k7_nbac``
"""

import hashlib
import io
import json
import re
import urllib.request
from datetime import date, datetime, timezone

import openpyxl

from etl.schema import RAW_DIR, SOURCES_JSON, STATUS_JSON

SOURCE_ID = "K7"
SERIES_BURNED_AREA = "nbac_annual_burned_area"
SERIES_FIRE_COUNT = "cnfdb_annual_fire_count"

# NBAC fører Canada som helhet, i tillegg til provinser og territorier. Vi tar
# inn landtotalen.
ENTITY = "CAN"

NBAC_DIR_URL = "https://cwfis.cfs.nrcan.gc.ca/downloads/nbac/"
NFDB_URL = (
    "https://cwfis.cfs.nrcan.gc.ca/downloads/nfdb/fire_pnt/current_version/"
    "NFDB_point_stats.xlsx"
)
LANDING_URL = "https://cwfis.cfs.nrcan.gc.ca/ha/nfdb"
AGREEMENT_URL = "https://cwfis.cfs.nrcan.gc.ca/downloads/nbac/NBAC_EULA.pdf"

RAW_NBAC = RAW_DIR / "k7_nbac_summarystats.xlsx"
RAW_NFDB = RAW_DIR / "k7_nfdb_point_stats.xlsx"

# Utgivelsesdatoen står sist i filnavnet, og er det vi sorterer på.
NBAC_FILNAVN = re.compile(r"NBAC_summarystats_(\d{4})_to_(\d{4})_(\d{8})\.xlsx")

NBAC_ARK = "sumstats_admin_name"
NBAC_KOLONNE = "CANADA"

NFDB_ARK = "NFDB_Summary_Stats"
NFDB_AAR = "YEAR"
NFDB_BRANNER = "FIRES"

# Nederst i NFDB-arket står gjennomsnittsrader («10 yr average») i årskolonnen.
# De er avledninger, ikke observasjoner, og skal ikke inn i datasettet.
AARSTALL = re.compile(r"^\d{4}$")

BRUKERAGENT = "skogbranner-etl/1.0 (+https://github.com/sheffielddivided/skogbranner)"


def _hent(url, timeout=300):
    forespoersel = urllib.request.Request(url, headers={"User-Agent": BRUKERAGENT})
    with urllib.request.urlopen(forespoersel, timeout=timeout) as svar:
        return svar.read()


def _sha256(data):
    return hashlib.sha256(data).hexdigest()


def finn_nyeste_nbac(katalog_html):
    """Velger den nyeste NBAC-aggregatfilen i katalogen.

    Returnerer (filnavn, utgivelsesdato). Reiser feil hvis katalogen ikke har
    en eneste fil på forventet form — da har produsenten endret navnemønster,
    og det skal stoppe kjøringen framfor å hente en tilfeldig fil.
    """
    treff = [
        (m.group(3), m.group(0)) for m in NBAC_FILNAVN.finditer(katalog_html)
    ]
    if not treff:
        raise ValueError(
            f"fant ingen fil på formen NBAC_summarystats_ÅÅÅÅ_to_ÅÅÅÅ_ÅÅÅÅMMDD.xlsx "
            f"i {NBAC_DIR_URL}"
        )
    dato, filnavn = max(treff)
    return filnavn, dato


def _finn_overskriftsrad(rader, paakrevd):
    """Finner raden som bærer kolonneoverskriftene.

    Arkene har tittel- og kildelinjer over selve tabellen, og antall slike
    linjer er ikke garantert å stå fast. Raden søkes derfor opp på innhold.
    """
    for i, rad in enumerate(rader):
        verdier = {str(c).strip() for c in rad if c is not None}
        if paakrevd <= verdier:
            return i
    raise ValueError(f"fant ingen rad med kolonnene {sorted(paakrevd)}")


def les_nbac(data):
    """Leser landtotalen for brent areal fra NBAC-aggregatet.

    Verdiene er justerte hektar. De returneres urørt — omregningen til km²
    skjer i normalize.py.
    """
    bok = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    if NBAC_ARK not in bok.sheetnames:
        raise ValueError(
            f"NBAC-filen mangler arket {NBAC_ARK!r}. Ark i filen: {bok.sheetnames}"
        )
    ark = bok[NBAC_ARK]
    rader = [rad for rad in ark.iter_rows(values_only=True)]

    i = _finn_overskriftsrad(rader, {NFDB_AAR, NBAC_KOLONNE})
    overskrifter = [str(c).strip() if c is not None else "" for c in rader[i]]
    kol_aar = overskrifter.index(NFDB_AAR)
    kol_canada = overskrifter.index(NBAC_KOLONNE)

    ut = []
    for rad in rader[i + 1 :]:
        if kol_aar >= len(rad) or kol_canada >= len(rad):
            continue
        aar, verdi = rad[kol_aar], rad[kol_canada]
        if aar is None or verdi is None:
            continue
        if not AARSTALL.match(str(aar).strip()):
            continue
        ut.append({"year": int(str(aar).strip()), "adjusted_ha": float(verdi)})

    if not ut:
        raise ValueError(f"arket {NBAC_ARK!r} ga ingen årsrader for {NBAC_KOLONNE}")
    ut.sort(key=lambda r: r["year"])
    return ut


def les_nfdb(data):
    """Leser antall registrerte branner per år for Canada som helhet."""
    bok = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    if NFDB_ARK not in bok.sheetnames:
        raise ValueError(
            f"NFDB-filen mangler arket {NFDB_ARK!r}. Ark i filen: {bok.sheetnames}"
        )
    ark = bok[NFDB_ARK]
    rader = [rad for rad in ark.iter_rows(values_only=True)]

    i = _finn_overskriftsrad(rader, {NFDB_AAR, NFDB_BRANNER})
    overskrifter = [str(c).strip() if c is not None else "" for c in rader[i]]
    kol_aar = overskrifter.index(NFDB_AAR)
    kol_branner = overskrifter.index(NFDB_BRANNER)

    ut = []
    for rad in rader[i + 1 :]:
        if kol_aar >= len(rad) or kol_branner >= len(rad):
            continue
        aar, antall = rad[kol_aar], rad[kol_branner]
        if aar is None or antall is None:
            continue
        # Gjennomsnittsradene nederst har tekst i årskolonnen.
        if not AARSTALL.match(str(aar).strip()):
            continue
        ut.append({"year": int(str(aar).strip()), "fires": int(round(float(antall)))})

    if not ut:
        raise ValueError(f"arket {NFDB_ARK!r} ga ingen årsrader")
    ut.sort(key=lambda r: r["year"])
    return ut


def hent():
    """Laster ned begge filene, skriver dem til data/raw/ og leser dem.

    Returnerer (areal_rader, brann_rader, info). Verdiene er hektar og antall
    slik de står i kilden.
    """
    RAW_DIR.mkdir(parents=True, exist_ok=True)

    katalog = _hent(NBAC_DIR_URL, timeout=120).decode("utf-8", errors="replace")
    filnavn, utgivelse = finn_nyeste_nbac(katalog)

    nbac_bytes = _hent(NBAC_DIR_URL + filnavn)
    RAW_NBAC.write_bytes(nbac_bytes)

    nfdb_bytes = _hent(NFDB_URL)
    RAW_NFDB.write_bytes(nfdb_bytes)

    areal = les_nbac(nbac_bytes)
    branner = les_nfdb(nfdb_bytes)

    info = {
        "source_id": SOURCE_ID,
        "series_id": [SERIES_BURNED_AREA, SERIES_FIRE_COUNT],
        "downloaded_at": date.today().isoformat(),
        "checksum": _sha256(nbac_bytes),
        "fire_count_checksum": _sha256(nfdb_bytes),
        "rows": len(areal) + len(branner),
        "nbac_file": filnavn,
        "nbac_release": f"{utgivelse[:4]}-{utgivelse[4:6]}-{utgivelse[6:]}",
        "coverage_start": min(areal[0]["year"], branner[0]["year"]),
        "coverage_end": max(areal[-1]["year"], branner[-1]["year"]),
    }
    return areal, branner, info


def skriv_metadata(info, processed_files):
    """Registrerer kilden i data/_sources.json.

    Siteringen og lenken til sluttbrukeravtalen gjengis ordrett. De er et
    vilkår for bruk, ikke en høflighet — se CLAUDE.md § 5.
    """
    with open(SOURCES_JSON, encoding="utf-8") as f:
        sources = json.load(f)

    sources["sources"][SOURCE_ID] = {
        "source_id": SOURCE_ID,
        "name": "CNFDB / NBAC — Canadian National Fire Database og National "
        "Burned Area Composite",
        "publisher": "Natural Resources Canada, Canadian Forest Service",
        "url": LANDING_URL,
        "download_url": NBAC_DIR_URL + info["nbac_file"],
        "fire_count_download_url": NFDB_URL,
        "license": "Krever aksept av sluttbrukeravtale. Avtalen er akseptert.",
        "license_url": AGREEMENT_URL,
        "attribution": (
            "Canadian Forest Service. 2021. Canadian National Fire Database – "
            "Agency Fire Data. Natural Resources Canada, Canadian Forest Service, "
            "Northern Forestry Centre, Edmonton, Alberta. "
            "https://cwfis.cfs.nrcan.gc.ca/ha/nfdb"
        ),
        "requires_agreement": True,
        "agreement_accepted": True,
        "geography": "Canada",
        "coverage_start": str(info["coverage_start"]),
        "coverage_end": str(info["coverage_end"]),
        "temporal_resolution": "annual",
        "quality": "reported",
        "unit_source": "hectares",
        "downloaded_at": info["downloaded_at"],
        "source_last_updated": info["nbac_release"],
        "checksum": info["checksum"],
        "fire_count_checksum": info["fire_count_checksum"],
        "series": [SERIES_BURNED_AREA, SERIES_FIRE_COUNT],
        "processed_files": processed_files,
        "footnotes": ["f_reporting_basis", "f_incomplete_inventory"],
        "notes": "Brent areal er produsentens eget årsaggregat fra National "
        "Burned Area Composite, oppgitt som justerte hektar. Antall branner "
        "kommer fra punktdataene i Canadian National Fire Database. Kilden "
        "opplyser selv at databasen verken er komplett eller feilfri, og at "
        "både fullstendighet og kvalitet varierer mellom byråene som "
        "rapporterer og mellom år.",
    }
    with open(SOURCES_JSON, "w", encoding="utf-8") as f:
        json.dump(sources, f, ensure_ascii=False, indent=2)
        f.write("\n")


def skriv_status(status, melding, info=None):
    """Skriver kjørestatus til data/_status.json."""
    with open(STATUS_JSON, encoding="utf-8") as f:
        data = json.load(f)

    naa = datetime.now(timezone.utc).isoformat(timespec="seconds")
    forrige = data["sources"].get(SOURCE_ID, {})
    data["last_run"] = naa
    data["sources"][SOURCE_ID] = {
        "status": status,
        "last_attempt": naa,
        "last_success": naa if status == "ok" else forrige.get("last_success"),
        "rows": info["rows"] if info else forrige.get("rows"),
        "checksum": info["checksum"] if info else forrige.get("checksum"),
        "message": melding,
    }
    with open(STATUS_JSON, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
        f.write("\n")


def main():
    areal, branner, info = hent()
    print(
        f"K7: {len(areal)} årsrader areal, {len(branner)} årsrader branntall, "
        f"{info['coverage_start']}–{info['coverage_end']} fra {info['nbac_file']}"
    )
    return areal, branner, info


if __name__ == "__main__":
    main()
